# Copyright (C) 2026 James Lander, The Thinking Home
# Licensed under GPL-3.0-or-later. See the LICENSE file in this repository.
# Device Sentinel - a Home Assistant custom integration from The Thinking Home (xeazy.com)
#   Article: https://xeazy.com/reliable-home-assistant-dead-sensor-detection/
#   Repository: https://github.com/TheThinkingHome/device_sentinel
# File: tools/simulation/scenario.py, Version: 0.20.5 (2026-09-05)

"""Read the simulation scenario workbook.

The workbook is the scenario script and lives outside this repository,
beside the fleet files, because it carries real device names from two
real houses. This module only knows how to read it, so the code is
public and the data is not.

Three tabs matter. Devices carries each device's rhythm, drawn from
its own learned statistics rather than invented. Scenario is a grid,
one row per device and one column per day, each cell naming what that
device does that day. System carries the outages, restarts and time
boundaries, each with the result expected of it, written down before
the run so it cannot be adjusted afterwards.

Codes are distinct and upper case on purpose. The first build of the
workbook used `X` and `x` for excluded and un-excluded, which double
counted 127 cells because COUNTIF ignores case, and would have let a
typo silently mean the opposite of what was intended.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

NORMAL = "."
FREEZE = "F"
UNAVAILABLE = "U"
NEVER = "N"
BATTERY_STEP = "B"
BATTERY_REPLACED = "R"
BATTERY_ZERO = "Z"
BATTERY_PLATEAU = "P"
BATTERY_NOISE = "J"
SIGNAL_BAD = "S"
SIGNAL_RAILED = "L"
EXCLUDED = "X"
UN_EXCLUDED = "UX"
MUTED = "M"
UN_MUTED = "UM"

CODES = {
    NORMAL, FREEZE, UNAVAILABLE, NEVER, BATTERY_STEP, BATTERY_REPLACED,
    BATTERY_ZERO, BATTERY_PLATEAU, BATTERY_NOISE, SIGNAL_BAD,
    SIGNAL_RAILED, EXCLUDED, UN_EXCLUDED, MUTED, UN_MUTED,
}

WORKBOOK_NAME = "device_sentinel_simulation_scenario.xlsx"


def workbook_path() -> Path:
    """Where the scenario lives: beside the fleet files."""
    root = os.environ.get("DEVICE_SENTINEL_FLEET_DIR", "/home/claude/fleets")
    return Path(root) / WORKBOOK_NAME


@dataclass
class Device:
    """One simulated device, carrying a real device's rhythm."""

    name: str
    fleet: str
    integration: str
    clock: str
    median_gap_s: float
    max_gap_s: float
    battery: float | None
    battery_rate_wk: float | None
    signal_median: float | None
    signal_floor: float | None
    rhythm_source: str
    daily_max_median_s: float | None = None
    observed_days: float | None = None
    events: int | None = None
    days: list[str] = field(default_factory=list)

    @property
    def has_battery(self) -> bool:
        return self.battery is not None

    @property
    def has_signal(self) -> bool:
        return self.signal_floor is not None

    def code(self, day: int) -> str:
        """The code for a one-based day number."""
        return self.days[day - 1]


@dataclass
class SystemEvent:
    """One thing done to the whole house, and what is expected of it."""

    day: int
    on: date
    kind: str
    scope: str
    detail: str
    expect: str


@dataclass
class Scenario:
    """The whole scenario: devices, days, and system events."""

    seed: int
    start: date
    days: int
    devices: list[Device]
    system: list[SystemEvent]

    def by_fleet(self, fleet: str) -> list[Device]:
        return [d for d in self.devices if d.fleet == fleet]

    def on_day(self, day: int, kind: str | None = None) -> list[SystemEvent]:
        return [
            e for e in self.system
            if e.day == day and (kind is None or e.kind == kind)
        ]

    def injected(self, day: int) -> list[tuple[Device, str]]:
        """Every device doing something other than normal on this day."""
        return [
            (d, d.code(day)) for d in self.devices
            if d.code(day) != NORMAL
        ]


def _cell(value, cast=None, default=None):
    if value is None or value == "":
        return default
    return cast(value) if cast else value


def load(path: Path | None = None) -> Scenario:
    """Read the workbook. Raises FileNotFoundError when it is absent,
    which every caller turns into a skip."""
    from openpyxl import load_workbook

    path = path or workbook_path()
    book = load_workbook(path, data_only=True, read_only=True)

    legend = {
        row[0]: row[1]
        for row in book["Legend"].iter_rows(min_row=1, max_col=2,
                                            values_only=True)
        if row and row[0]
    }
    seed = int(legend["Seed"])
    start = date.fromisoformat(str(legend["Day 1"])[:10])

    grid = book["Scenario"]
    header = [c.value for c in next(grid.iter_rows(min_row=1, max_row=1))]
    day_count = sum(1 for h in header if isinstance(h, str)
                    and h.startswith("D") and h[1:].isdigit())

    profiles = {}
    for row in book["Devices"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        profiles[row[0]] = row

    devices: list[Device] = []
    for row in grid.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = row[0]
        p = profiles.get(name)
        if p is None:
            raise ValueError(f"{name} is on Scenario but not on Devices")
        codes = [str(c) if c is not None else NORMAL
                 for c in row[3:3 + day_count]]
        bad = {c for c in codes} - CODES
        if bad:
            raise ValueError(f"{name} carries unknown code(s): {sorted(bad)}")
        devices.append(Device(
            name=name, fleet=p[1], integration=p[2], clock=p[3],
            median_gap_s=float(p[4]), max_gap_s=float(p[5]),
            battery=_cell(p[6], float), battery_rate_wk=_cell(p[7], float),
            signal_median=_cell(p[8], float), signal_floor=_cell(p[9], float),
            rhythm_source=p[10],
            daily_max_median_s=_cell(p[11], float),
            observed_days=_cell(p[12], float),
            events=_cell(p[13], int),
            days=codes,
        ))

    system = []
    for row in book["System"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        system.append(SystemEvent(
            day=int(row[0]), on=date.fromisoformat(str(row[1])[:10]),
            kind=row[2], scope=row[3], detail=row[4], expect=row[5],
        ))

    book.close()
    return Scenario(seed=seed, start=start, days=day_count,
                    devices=devices, system=system)
